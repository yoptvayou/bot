# --- Импорты ---
from telegram.constants import ParseMode
import atexit
import logging
import re
import os
import base64
import json
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Set
from collections import defaultdict, deque
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import openpyxl # type: ignore
import warnings
import sys
import io
import asyncio
from concurrent.futures import ThreadPoolExecutor

# --- Подавление предупреждений от openpyxl ---
warnings.filterwarnings("ignore", message="Data Validation extension is not supported", category=UserWarning)

# --- Настройка логирования ---
# Конфигурация логирования для отслеживания действий бота
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO # Можно установить DEBUG для более подробных логов
)
logging.getLogger("httpx").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# --- Конфигурация ---
# Константа для города, используется в поиске файлов
CITY = 'Воронеж'
# Разрешения для Google Drive API
SCOPES = ['https://www.googleapis.com/auth/drive']
# Путь к директории для хранения локальных кэшированных файлов
LOCAL_CACHE_DIR = "./local_cache"
# --- Глобальные переменные ---
# Путь к файлу учетных данных Google
CREDENTIALS_FILE: str = ""
# Токен Telegram бота
TELEGRAM_TOKEN: str = ""
# ID родительской папки в Google Drive
PARENT_FOLDER_ID: str = ""
# ID временной папки в Google Drive
TEMP_FOLDER_ID: str = ""
# ID корневой папки года в Google Drive
ROOT_FOLDER_YEAR: str = ""
# ID файла черного списка пользователей
BLACKLIST_FILE_ID: str = ""
# ID файла белого списка пользователей
WHITELIST_FILE_ID: str = ""
# Часовой пояс (по умолчанию UTC)
TIMEZONE_OFFSET = int(os.getenv("TIMEZONE_OFFSET", 3))  # Часы от UTC (например, 3 для MSK)
# ID последнего загруженного файла
LAST_FILE_ID: Optional[str] = None
# Дата последнего файла
LAST_FILE_DATE: Optional[datetime] = None
# Время последнего изменения файла в Google Drive
LAST_FILE_DRIVE_TIME: Optional[datetime] = None
# Локальный путь к последнему файлу
LAST_FILE_LOCAL_PATH: Optional[str] = None
# Пул потоков для параллельной обработки
executor = ThreadPoolExecutor(max_workers=4)

# --- Разрешённые пользователи (администраторы) ---
# Список пользователей с правами администратора
ALLOWED_USERS = {'tupikin_ik', 'yoptvayou'}

# --- Защита от DDoS ---
# Лимиты сообщений (количество сообщений за период)
MESSAGE_LIMITS = {
    'minute': 10,   # 10 сообщений в минуту
    'hour': 100,    # 100 сообщений в час
    'day': 1000     # 1000 сообщений в день
}

# Хранилище для отслеживания активности пользователей
user_activity: Dict[str, Dict[str, deque]] = defaultdict(lambda: {
    'minute': deque(),
    'hour': deque(),
    'day': deque()
})

# Блокировка пользователей (черный список)
banned_users: Set[str] = set()
# Время блокировки пользователей (в минутах)
user_ban_times: Dict[str, int] = {}
# Время начала блокировки
user_ban_start_times: Dict[str, datetime] = {}

# --- Функции для работы с учетными данными ---
def get_credentials_path() -> str:
    """
    Декодирует Google Credentials из переменной окружения.
    Returns:
        str: Путь к временному файлу с учетными данными
    Raises:
        RuntimeError: Если переменная окружения GOOGLE_CREDS_BASE64 не найдена
    """
    # Получаем закодированные учетные данные из переменной окружения
    encoded = os.getenv("GOOGLE_CREDS_BASE64")
    if not encoded:
        raise RuntimeError("GOOGLE_CREDS_BASE64 не найдена!")
    try:
        # Расшифровываем данные и сохраняем во временный файл
        decoded = base64.b64decode(encoded).decode('utf-8')
        creds = json.loads(decoded)
        temp_path = "temp_google_creds.json"
        with open(temp_path, 'w') as f:
            json.dump(creds, f)
        logger.info(f"✅ Учетные данные сохранены: {temp_path}")
        # Регистрируем функцию для удаления временного файла при выходе
        atexit.register(lambda: os.remove(temp_path) if os.path.exists(temp_path) else None)
        return temp_path
    except Exception as e:
        logger.error(f"❌ Ошибка декодирования GOOGLE_CREDS_BASE64: {e}")
        raise

def init_config():
    """
    Инициализация конфигурации бота из переменных окружения.
    Raises:
        RuntimeError: Если не все необходимые переменные окружения установлены
    """
    global CREDENTIALS_FILE, TELEGRAM_TOKEN, PARENT_FOLDER_ID, TEMP_FOLDER_ID, ROOT_FOLDER_YEAR, BLACKLIST_FILE_ID, WHITELIST_FILE_ID, TIMEZONE_OFFSET
    # Получаем путь к учетным данным
    CREDENTIALS_FILE = get_credentials_path()
    # Получаем токен Telegram бота
    TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
    # Получаем ID родительской папки
    PARENT_FOLDER_ID = os.getenv("PARENT_FOLDER_ID", "")
    # Получаем ID временной папки
    TEMP_FOLDER_ID = os.getenv("TEMP_FOLDER_ID", "")
    # Получаем ID файлов черного и белого списков
    BLACKLIST_FILE_ID = os.getenv("BLACKLIST_FILE_ID", "")
    WHITELIST_FILE_ID = os.getenv("WHITELIST_FILE_ID", "")
    # Получаем часовой пояс
    TIMEZONE_OFFSET = int(os.getenv("TIMEZONE_OFFSET", 3))
    # Устанавливаем год для корневой папки
    ROOT_FOLDER_YEAR = str(datetime.now().year)
    # Проверяем наличие всех необходимых переменных
    if not TELEGRAM_TOKEN or not PARENT_FOLDER_ID or not BLACKLIST_FILE_ID or not WHITELIST_FILE_ID:
        missing = []
        if not TELEGRAM_TOKEN: missing.append("TELEGRAM_TOKEN")
        if not PARENT_FOLDER_ID: missing.append("PARENT_FOLDER_ID")
        if not BLACKLIST_FILE_ID: missing.append("BLACKLIST_FILE_ID")
        if not WHITELIST_FILE_ID: missing.append("WHITELIST_FILE_ID")
        raise RuntimeError(f"❌ Отсутствуют переменные окружения: {', '.join(missing)}")
    # Создаем директорию для кэширования
    os.makedirs(LOCAL_CACHE_DIR, exist_ok=True)
    logger.info(f"📁 Локальный кэш: {os.path.abspath(LOCAL_CACHE_DIR)}")

# --- Класс для работы с Google API ---
class GoogleServices:
    """
    Singleton класс для работы с Google Drive API.
    Этот класс обеспечивает единственный экземпляр соединения с Google Drive API,
    что позволяет избежать многократного создания соединений.
    """
    # Статический атрибут для хранения экземпляра класса
    _instance = None
    def __new__(cls):
        """
        Переопределение метода __new__ для реализации паттерна Singleton.
        Returns:
            GoogleServices: Единственный экземпляр класса
        """
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            # Создаем учетные данные из файла
            creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
            # Инициализируем сервис Google Drive
            cls._instance.drive = build('drive', 'v3', credentials=creds)
        return cls._instance

# --- Класс управления доступом ---
class AccessManager:
    """
    Класс для управления доступом пользователей через черный и белый списки.
    Осуществляет проверку доступа пользователей к функционалу бота.
    """
    def __init__(self, drive_service):
        """
        Инициализация менеджера доступа.
        Args:
            drive_service: Сервис Google Drive для работы с файлами
        """
        self.drive = drive_service
        self.blacklist = set()
        self.whitelist = set()
    def download_list(self, file_id: str) -> List[str]:
        """
        Скачивает список пользователей из Google Drive файла.
        Args:
            file_id (str): ID файла в Google Drive
        Returns:
            List[str]: Список username пользователей (без @, в нижнем регистре)
        """
        try:
            # Получаем медиа-поток файла
            request = self.drive.files().get_media(fileId=file_id)
            file_data = io.BytesIO()
            downloader = MediaIoBaseDownload(file_data, request)
            done = False
            # Скачиваем файл
            while not done:
                status, done = downloader.next_chunk()
            file_data.seek(0)
            content = file_data.read().decode('utf-8')
            # Обрабатываем содержимое файла
            usernames = []
            for line in content.splitlines():
                # Очищаем строку: удаляем @, приводим к нижнему регистру, убираем пробелы
                cleaned = line.strip().lower().replace('@', '')
                if cleaned:
                    usernames.append(cleaned)
            return usernames
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки списка из файла {file_id}: {e}")
            return []

    def update_lists(self):
        """
        Обновляет черный и белый списки пользователей из Google Drive файлов.
        """
        # Загружаем белый список
        if WHITELIST_FILE_ID:
            self.whitelist = set(self.download_list(WHITELIST_FILE_ID))
            logger.info(f"✅ Загружен белый список: {len(self.whitelist)} пользователей")
        else:
            logger.warning("⚠️ WHITELIST_FILE_ID не задан — белый список пуст")
        # Загружаем черный список
        if BLACKLIST_FILE_ID:
            self.blacklist = set(self.download_list(BLACKLIST_FILE_ID))
            logger.info(f"✅ Загружен чёрный список: {len(self.blacklist)} пользователей")
        else:
            logger.warning("⚠️ BLACKLIST_FILE_ID не задан — чёрный список пуст")

    def is_allowed(self, username: str) -> bool:
        """
        Проверяет, разрешен ли доступ пользователю.
        Args:
            username (str): Имя пользователя Telegram
        Returns:
            bool: True, если доступ разрешен, False в противном случае
        """
        if not username:
            return False
        username_lower = username.lower()
        # Администраторы всегда имеют доступ
        if username_lower in {u.lower() for u in ALLOWED_USERS}:
            return True
        # Чёрный список — запрещает доступ, даже если в белом
        if username_lower in self.blacklist:
            return False
        # Если белый список активен — только он определяет доступ
        if self.whitelist and username_lower not in self.whitelist:
            return False
        # Если белый список пуст — разрешаем всех, кроме чёрного
        return True

# Глобальная переменная для менеджера доступа
access_manager: Optional[AccessManager] = None

# --- Функции защиты от DDoS ---
def check_user_limit(username: str) -> bool:
    """
    Проверяет, превышает ли пользователь лимиты сообщений.
    Args:
        username (str): Имя пользователя Telegram
    Returns:
        bool: True, если пользователь не заблокирован и лимиты не превышены
    """
    # Проверяем, заблокирован ли пользователь
    if username in banned_users:
        # Проверяем, истекло ли время блокировки
        if username in user_ban_start_times:
            ban_duration = timedelta(minutes=user_ban_times.get(username, 10))
            ban_start = user_ban_start_times[username]
            if datetime.now(timezone.utc) + timedelta(hours=TIMEZONE_OFFSET) >= ban_start + ban_duration:
                # Время блокировки истекло, разблокируем пользователя
                unban_user(username)
                logger.info(f"🔓 Пользователь {username} разблокирован автоматически")
                # Удаляем информацию о блокировке
                user_ban_start_times.pop(username, None)
                user_ban_times.pop(username, None)
                return True
            else:
                # Пользователь всё ещё заблокирован, выводим время до разблокировки
                remaining_time = ban_start + ban_duration - (datetime.now(timezone.utc) + timedelta(hours=TIMEZONE_OFFSET))
                minutes_left = int(remaining_time.total_seconds() // 60)
                logger.warning(f"⚠️ Пользователь {username} заблокирован. Осталось {minutes_left} минут")
                return False
        else:
            # Время блокировки не указано, разблокируем
            unban_user(username)
            return True

    now = datetime.now(timezone.utc) + timedelta(hours=TIMEZONE_OFFSET)
    # Очищаем устаревшие записи
    for period, queue in user_activity[username].items():
        # Исправленный код для timedelta
        time_delta_key = period + 's'  # minute -> minutes, hour -> hours, day -> days
        if time_delta_key in ['minutes', 'hours', 'days']:
            delta = timedelta(**{time_delta_key: 1})
            while queue and queue[0] <= now - delta:
                queue.popleft()
        else:
            logger.warning(f"⚠️ Неподдерживаемый период: {period}")

    # Проверяем лимиты
    for period, limit in MESSAGE_LIMITS.items():
        if len(user_activity[username][period]) >= limit:
            logger.warning(f"⚠️ Пользователь {username} превысил лимит {limit} сообщений за {period}")
            ban_user(username)
            return False
    # Добавляем текущее сообщение
    for period in MESSAGE_LIMITS.keys():
        user_activity[username][period].append(now)
    return True

def ban_user(username: str):
    """
    Блокирует пользователя.
    Args:
        username (str): Имя пользователя Telegram
    """
    # Определяем время блокировки (начинается с 10 минут, увеличивается на 10 каждые 10 минут)
    ban_time = user_ban_times.get(username, 10)
    user_ban_times[username] = ban_time + 10
    user_ban_start_times[username] = datetime.now(timezone.utc) + timedelta(hours=TIMEZONE_OFFSET)
    banned_users.add(username)
    logger.info(f"🔒 Пользователь {username} заблокирован на {ban_time} минут")

def unban_user(username: str):
    """
    Разблокирует пользователя.
    Args:
        username (str): Имя пользователя Telegram
    """
    banned_users.discard(username)
    logger.info(f"🔓 Пользователь {username} разблокирован")
    # Удаляем информацию о блокировке
    user_ban_start_times.pop(username, None)
    user_ban_times.pop(username, None)

def reset_user_limits(username: str):
    """
    Сбрасывает лимиты для пользователя.
    Args:
        username (str): Имя пользователя Telegram
    """
    if username in user_activity:
        for period in MESSAGE_LIMITS.keys():
            user_activity[username][period].clear()
    logger.info(f"🔄 Лимиты для пользователя {username} сброшены")
    # Сбрасываем информацию о блокировке
    user_ban_start_times.pop(username, None)
    user_ban_times.pop(username, None)

# --- Команда /whitelist ---
async def manage_whitelist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Управление белым списком: добавить, удалить, показать.
    Доступно только администраторам.
    Использование:
      /whitelist show
      /whitelist add @username1 @username2
      /whitelist remove @username1 @username2
    """
    if not update.message or not update.effective_user:
        return
    user = update.effective_user
    if not user.username or user.username.lower() not in {u.lower() for u in ALLOWED_USERS}:
        await update.message.reply_text(get_message('admin_only'))
        return

    if not access_manager:
        await update.message.reply_text("❌ Система доступа не инициализирована.")
        return

    args = context.args
    if not args:
        await update.message.reply_text(
            get_message('list_usage', list_type='whitelist'),
            parse_mode='HTML'
        )
        return

    action = args[0].lower()
    usernames = [u.lstrip('@').lower() for u in args[1:]] if len(args) > 1 else []

    if action == "show":
        if access_manager.whitelist:
            whitelist_text = "\n".join([f"@{u}" for u in sorted(access_manager.whitelist)])
            await update.message.reply_text(
                get_message('list_show_header',
                           list_type='Белый',
                           count=len(access_manager.whitelist),
                           usernames=whitelist_text),
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                get_message('list_show_empty', list_type='Белый')
            )
        return

    elif action in ["add", "remove"]:
        if not usernames:
            await update.message.reply_text(
                get_message('list_no_usernames')
            )
            return

        # Проверка разрешений на запись в Google Drive перед изменением
        gs = GoogleServices()
        fm = FileManager(gs.drive)
        can_write_whitelist = fm.check_write_permission(WHITELIST_FILE_ID)
        can_write_blacklist = fm.check_write_permission(BLACKLIST_FILE_ID)

        if not (can_write_whitelist and can_write_blacklist):
            await update.message.reply_text(
                get_message('list_no_write_permission', list_type='списков')
            )
            logger.warning(f"Администратор {user.username} попытался изменить списки, но у бота нет прав на запись.")
            return

        if action == "add":
            added = []
            already_in = []
            for u in usernames:
                if u not in access_manager.whitelist:
                    access_manager.whitelist.add(u)
                    added.append(u)
                else:
                    already_in.append(u)
            
            # Обновляем файл на Google Drive
            success = fm.update_list_file(WHITELIST_FILE_ID, sorted(access_manager.whitelist))
            if success:
                msg_added = ', '.join([f'@{u}' for u in added]) if added else "—"
                msg_already = ', '.join([f'@{u}' for u in already_in]) if already_in else "—"
                await update.message.reply_text(
                    get_message('list_update_success_add',
                               list_type='Белый',
                               added=msg_added,
                               already_in=msg_already)
                )
                logger.info(f"Администратор {user.username} добавил в белый список: {added}")
            else:
                # Откатываем изменения в памяти, если запись не удалась
                for u in added:
                    access_manager.whitelist.discard(u)
                await update.message.reply_text(
                    get_message('list_update_error', list_type='белого списка')
                )

        elif action == "remove":
            removed = []
            not_found = []
            for u in usernames:
                if u in access_manager.whitelist:
                    access_manager.whitelist.discard(u)
                    removed.append(u)
                else:
                    not_found.append(u)
            
            # Обновляем файл на Google Drive
            success = fm.update_list_file(WHITELIST_FILE_ID, sorted(access_manager.whitelist))
            if success:
                msg_removed = ', '.join([f'@{u}' for u in removed]) if removed else "—"
                msg_not_found = ', '.join([f'@{u}' for u in not_found]) if not_found else "—"
                await update.message.reply_text(
                    get_message('list_update_success_remove',
                               list_type='Белый',
                               removed=msg_removed,
                               not_found=msg_not_found)
                )
                logger.info(f"Администратор {user.username} удалил из белого списка: {removed}")
            else:
                # Откатываем изменения в памяти, если запись не удалась
                for u in removed:
                    access_manager.whitelist.add(u)
                await update.message.reply_text(
                    get_message('list_update_error', list_type='белого списка')
                )
    else:
        await update.message.reply_text(
            get_message('list_unknown_action'),
            parse_mode='HTML'
        )


async def manage_blacklist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Управление чёрным списком: добавить, удалить, показать.
    Доступно только администраторам.
    Использование:
      /blacklist show
      /blacklist add @username1 @username2
      /blacklist remove @username1 @username2
    """
    if not update.message or not update.effective_user:
        return
    user = update.effective_user
    if not user.username or user.username.lower() not in {u.lower() for u in ALLOWED_USERS}:
        await update.message.reply_text(get_message('admin_only'))
        return

    if not access_manager:
        await update.message.reply_text("❌ Система доступа не инициализирована.")
        return

    args = context.args
    if not args:
        await update.message.reply_text(
            get_message('list_usage', list_type='blacklist'),
            parse_mode='HTML'
        )
        return

    action = args[0].lower()
    usernames = [u.lstrip('@').lower() for u in args[1:]] if len(args) > 1 else []

    if action == "show":
        if access_manager.blacklist:
            blacklist_text = "\n".join([f"@{u}" for u in sorted(access_manager.blacklist)])
            await update.message.reply_text(
                get_message('list_show_header',
                           list_type='Чёрный',
                           count=len(access_manager.blacklist),
                           usernames=blacklist_text),
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                get_message('list_show_empty', list_type='Чёрный')
            )
        return

    elif action in ["add", "remove"]:
        if not usernames:
            await update.message.reply_text(
                get_message('list_no_usernames')
            )
            return

        # Проверка разрешений на запись в Google Drive перед изменением
        gs = GoogleServices()
        fm = FileManager(gs.drive)
        can_write_whitelist = fm.check_write_permission(WHITELIST_FILE_ID)
        can_write_blacklist = fm.check_write_permission(BLACKLIST_FILE_ID)

        if not (can_write_whitelist and can_write_blacklist):
             await update.message.reply_text(
                get_message('list_no_write_permission', list_type='списков')
            )
        logger.warning(f"Администратор {user.username} попытался изменить списки, но у бота нет прав на запись.")
        return

    if action == "add":
        added = []
        already_in = []
        for u in usernames:
            if u not in access_manager.blacklist:
                access_manager.blacklist.add(u)
                # Автоматически удаляем из белого списка при добавлении в чёрный
                if u in access_manager.whitelist:
                    access_manager.whitelist.discard(u)
                added.append(u)
            else:
                already_in.append(u)
            
        # Обновляем файлы на Google Drive
        success_black = fm.update_list_file(BLACKLIST_FILE_ID, sorted(access_manager.blacklist))
        success_white = fm.update_list_file(WHITELIST_FILE_ID, sorted(access_manager.whitelist)) # Обновляем белый список тоже
            
        if success_black and success_white:
            msg_added = ', '.join([f'@{u}' for u in added]) if added else "—"
            msg_already = ', '.join([f'@{u}' for u in already_in]) if already_in else "—"
            await update.message.reply_text(
                    get_message('list_update_success_add',
                            list_type='Чёрный',
                            added=msg_added,
                            already_in=msg_already)
            )
            logger.info(f"Администратор {user.username} добавил в чёрный список: {added}")
        else:
            # Откатываем изменения в памяти, если запись не удалась
            for u in added:
                access_manager.blacklist.discard(u)
                # Восстанавливаем в белый список, если был удален
                # (Логика восстановления может быть сложнее, опущена для простоты)
            await update.message.reply_text(
                get_message('list_update_error', list_type='чёрного списка')
            )

    elif action == "remove":
        removed = []
        not_found = []
        for u in usernames:
            if u in access_manager.blacklist:
                access_manager.blacklist.discard(u)
                removed.append(u)
            else:
                not_found.append(u)
            
        # Обновляем файл на Google Drive
        success = fm.update_list_file(BLACKLIST_FILE_ID, sorted(access_manager.blacklist))
        if success:
            msg_removed = ', '.join([f'@{u}' for u in removed]) if removed else "—"
            msg_not_found = ', '.join([f'@{u}' for u in not_found]) if not_found else "—"
            await update.message.reply_text(
                    get_message('list_update_success_remove',
                            list_type='Чёрный',
                            removed=msg_removed,
                            not_found=msg_not_found)
            )
            logger.info(f"Администратор {user.username} удалил из чёрного списка: {removed}")
        else:
            # Откатываем изменения в памяти, если запись не удалась
            for u in removed:
                access_manager.blacklist.add(u)
            await update.message.reply_text(
                get_message('list_update_error', list_type='чёрного списка')
            )
    else:
        await update.message.reply_text(
            get_message('list_unknown_action'),
            parse_mode='HTML'
        )

# --- Ответы бота ---
def get_message(message_code: str, **kwargs) -> str:
    """
    Возвращает текст сообщения по коду с возможностью подстановки параметров.
    Args:
        message_code (str): Код сообщения
        **kwargs: Параметры для подстановки в шаблон
    Returns:
        str: Форматированное сообщение
    """
    messages = {
        'access_denied': (
            "Ты кто такой, дядя?\n"
            "Не в списке — не входи.\n"
            "Хочешь доступ — плати бабки или лежи в багажнике до утра."
        ),
        'help': (
            "О, смотри-ка — гость на складе!\n"
            "Только не стой как лох у контейнера — говори, что надо.\n"
            "• <code>/s 123456</code> — найти терминал по СН\n"
            "• <code>/s 123456, 789012</code> — найти несколько терминалов по СН\n"            
            "• <code>@Sklad_bot 123456</code> — крикни в чатике, я найду\n"
            "\n"
            "<b>Только для админов:</b>\n"
            "• <code>/whitelist show|add|remove [@username...]</code> — управление белым списком\n"
            "• <code>/blacklist show|add|remove [@username...]</code> — управление чёрным списком\n"
            "• <code>/path</code> — глянуть, что у нас в папке завалялось\n"
            "• <code>/reload_lists</code> — обновить список предателей и своих\n"
            "• <code>/restart</code> — перезапуск бота\n"
            "• <code>/refresh</code> — обновления файла склада\n"
            "• <code>/reset_bans</code> — сброс банов\n"
        ),
        'invalid_number': (
            "Ты чё, братан, по пьяни печатаешь?\n"
            "СН — это типа <code>AB123456</code>, без пробелов, без носков в клавиатуре.\n"
            "Попробуй ещё раз, а то выкину в реку."
        ),
        'search_start': (
            "🔍 Копаю в архивах... Где-то был этот <code>{number}</code>...\n"
            "Если не спёрли, как в прошлый раз — найду."
        ),
        'no_file': (
            "Архивы пусты, брат.\n"
            "Либо файл сожгли, либо его ещё не подкинули.\n"
            "Приходи завтра — может, кто-нибудь не сдохнет и загрузит."
        ),
        'file_not_found_local': (
            "Файл был, но теперь его нет.\n"
            "Кто-то слил базу в канализацию или сервер сдох.\n"
            "Жди, пока кто-то перезальёт."
        ),
        'no_terminal': (
            "Терминал с СН <code>{number}</code>?\n"
            "Нету. Ни в базе, ни в подвале, ни в багажнике 'Весты'.\n"
            "Может, он уже в металлоломе... или ты втираешь мне очки?"
        ),
        'file_update_error': (
            "Файл обновился, но я не смог его подтянуть.\n"
            "Работаю на старых данных — могут быть косяки."
        ),
        'file_update_success': (
            "Файл обновился, но я не смог его загрузить.\n"
            "Продолжаю работать на старых данных."
        ),
        'search_error': (
            "База есть, но читать не могу — видимо, кто-то опять говнокод написал.\n"
            "Попробуй позже."
        ),
        'missing_number': (
            "Укажи серийный номер после команды.\n"
            "Пример: <code>/s AB123456</code>"
        ),
        'unknown_command': (
            "Неизвестная команда.\n"
            "Доступные команды:\n"
            "• <code>/s 123456</code> — найти терминал по СН\n"
            "• <code>/s 123456, 789012</code> — найти несколько терминалов по СН\n"            
            "• <code>@Sklad_bot 123456</code> — крикни в чатике, я найду\n"
            "\n"
            "<b>Только для админов:</b>\n"
            "• <code>/whitelist show|add|remove [@username...]</code> — управление белым списком\n"
            "• <code>/blacklist show|add|remove [@username...]</code> — управление чёрным списком\n"
            "• <code>/path</code> — глянуть, что у нас в папке завалялось\n"
            "• <code>/reload_lists</code> — обновить список предателей и своих\n"
            "• <code>/restart</code> — перезапуск бота\n"
            "• <code>/refresh</code> — обновления файла склада\n"
            "• <code>/reset_bans</code> — сброс банов\n"
        ),
        'ddos_blocked': (
            "Ты слишком быстро пишешь! Тебе нужно немного передышки.\n"
            "Пожалуйста, подожди немного и попробуй снова."
        ),
        'reset_success': (
            "✅ Лимиты для пользователя <code>{username}</code> были сброшены."
        ),
        'reset_all_success': (
            "✅ Все лимиты были сброшены."
        ),
        'reset_fail': (
            "❌ Не удалось сбросить лимиты для пользователя <code>{username}</code>."
        ),
        'admin_only': (
            "❌ Эта команда доступна только администраторам."
        ),
        'list_show_empty': (
            "{list_type} список пуст."
        ),
        'list_show_header': (
            "<b>{list_type} список ({count}):</b>\n<code>{usernames}</code>"
        ),
        'list_usage': (
            "Использование:\n"
            "<code>/{list_type} show</code> — показать список\n"
            "<code>/{list_type} add @username1 @username2</code> — добавить пользователей\n"
            "<code>/{list_type} remove @username1 @username2</code> — удалить пользователей"
        ),
        'list_no_usernames': (
            "Укажите хотя бы один username."
        ),
        'list_no_write_permission': (
            "❌ Недостаточно прав для записи в файл {list_type} на Google Drive. Изменения не сохранены."
        ),
        'list_update_success_add': (
            "✅ {list_type} список обновлён.\n"
            "Добавлены: {added}\n"
            "Уже в списке: {already_in}"
        ),
        'list_update_success_remove': (
            "✅ {list_type} список обновлён.\n"
            "Удалены: {removed}\n"
            "Не найдены в списке: {not_found}"
        ),
        'list_update_error': (
            "❌ Ошибка при обновлении файла {list_type} на Google Drive. Изменения отменены."
        ),
        'list_unknown_action': (
            "Неизвестное действие. Используйте <code>show</code>, <code>add</code> или <code>remove</code>."
        )
    }
    # Получаем сообщение по коду и форматируем его с параметрами
    message = messages.get(message_code, "Неизвестное сообщение")
    return message.format(**kwargs) if kwargs else message

def preload_latest_file():
    """
    При старте бота ищет и загружает последний файл из архива.
    Ищет файл за последние 30 дней, начиная с сегодняшней даты.
    """
    global LAST_FILE_ID, LAST_FILE_DATE, LAST_FILE_DRIVE_TIME, LAST_FILE_LOCAL_PATH
    gs = GoogleServices()
    fm = FileManager(gs.drive)
    today = datetime.now()
    logger.info("🔍 Поиск последнего файла при старте бота...")
    # Проверяем файлы за последние 30 дней
    for days_back in range(31):
        target_date = today - timedelta(days=days_back)
        filename = f"АПП_Склад_{target_date.strftime('%d%m%y')}_{CITY}.xlsm"
        # Ищем папку "акты"
        acts = fm.find_folder(PARENT_FOLDER_ID, "акты")
        if not acts:
            continue
        # Формируем имя месяца
        month_num = target_date.month
        month_name = ["январь", "февраль", "март", "апрель", "май", "июнь",
                      "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"][month_num - 1]
        month_folder = fm.find_folder(acts, f"{target_date.strftime('%m')} - {month_name}")
        if not month_folder:
            continue
        # Ищем папку с датой
        date_folder = fm.find_folder(month_folder, target_date.strftime('%d%m%y'))
        if not date_folder:
            continue
        # Ищем файл
        file_id = fm.find_file(date_folder, filename)
        if file_id:
            drive_time = fm.get_file_modified_time(file_id)
            if not drive_time:
                continue
            # Формируем локальный путь
            local_path = os.path.join(LOCAL_CACHE_DIR, f"cache_{target_date.strftime('%Y%m%d')}.xlsm")
            # Проверяем, нуждается ли файл в обновлении
            download_needed = True
            if os.path.exists(local_path):
                local_time = datetime.fromtimestamp(os.path.getmtime(local_path), tz=timezone.utc)
                if drive_time <= local_time:
                    download_needed = False
            # Скачиваем файл при необходимости
            if download_needed:
                logger.info(f"📥 Скачивание файла при старте: {filename} → {local_path}")
                if not fm.download_file(file_id, local_path):
                    logger.error("❌ Не удалось скачать файл при старте.")
                    continue
                logger.info(f"✅ Файл успешно загружен при старте: {local_path}")
            else:
                logger.info(f"✅ Используем существующий кэш: {local_path}")
            # Сохраняем метаданные файла
            LAST_FILE_ID = file_id
            LAST_FILE_DATE = target_date
            LAST_FILE_DRIVE_TIME = drive_time
            LAST_FILE_LOCAL_PATH = local_path
            logger.info(f"📁 Предзагружен файл: {filename} (ID: {file_id}) от {target_date.strftime('%d.%m.%Y')}")
            return
    # Если не нашли файл за 30 дней
    logger.warning("⚠️ Не удалось найти актуальный файл при старте.")
    LAST_FILE_ID = None
    LAST_FILE_DATE = None
    LAST_FILE_DRIVE_TIME = None
    LAST_FILE_LOCAL_PATH = None

def extract_number(query: str) -> Optional[str]:
    """
    Извлекает серийный номер из строки запроса.
    Args:
        query (str): Строка запроса пользователя
    Returns:
        Optional[str]: Очищенный серийный номер или None
    """
    if not query:
        return None
    # Удаляем все пробелы и лишние символы
    clean = re.sub(r'[^A-Za-z0-9\-]', '', query.strip())
    # Проверяем, соответствует ли строка формату СН
    if clean and re.fullmatch(r'[A-Za-z0-9\-]+', clean):
        return clean.upper()  # Приводим к верхнему регистру для единообразия
    return None

# --- Обработчики команд ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обработчик команды /start.
    Args:
        update (Update): Объект обновления от Telegram
        context (ContextTypes.DEFAULT_TYPE): Контекст обработчика
    """
    if not update.message:
        return
    user = update.effective_user
    chat_type = update.message.chat.type
    # Проверяем доступ в приватном чате
    if chat_type == 'private' and (not user.username or user.username not in ALLOWED_USERS):
        await update.message.reply_text(get_message('access_denied'))
        return
    await update.message.reply_text(get_message('help'), parse_mode='HTML')

# Обработчик команды /restart ---
async def restart_bot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Перезапуск бота (только для админов).
    Args:
        update (Update): Объект обновления от Telegram
        context (ContextTypes.DEFAULT_TYPE): Контекст обработчика
    """
    if not update.message or not update.effective_user:
        return
    user = update.effective_user
    # Проверяем, является ли пользователь администратором
    if not user.username or user.username.lower() not in {u.lower() for u in ALLOWED_USERS}:
        await update.message.reply_text("❌ Доступ запрещён.")
        return
    try:
        await update.message.reply_text("🔄 Перезапуск бота...")
        logger.info(f"🔄 Администратор {user.username} запустил перезагрузку бота.")
        # Используем os.execv для перезапуска текущего процесса
        os.execv(sys.executable, [sys.executable] + sys.argv)
        await update.message.reply_text("✅ Бот успешно перезагружен!")
    except Exception as e:
        logger.error(f"❌ Ошибка при перезапуске бота: {e}")
        await update.message.reply_text("❌ Произошла ошибка при перезагрузке бота.")

async def show_path(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Показать содержимое папки в Google Drive.
    Args:
        update (Update): Объект обновления от Telegram
        context (ContextTypes.DEFAULT_TYPE): Контекст обработчика
    """
    if update.message.chat.type == 'private':
        user = update.effective_user
        if not user.username or not access_manager.is_allowed(user.username):
            await update.message.reply_text(
                get_message('access_denied')
            )
            return
    try:
        gs = GoogleServices()
        fm = FileManager(gs.drive)
        root_id = PARENT_FOLDER_ID
        items = fm.list_files_in_folder(root_id, max_results=100)
        text = f"🗂 <b>Корневая папка</b> (ID: <code>{root_id}</code>)"
        # Формируем текст ответа
        if not items:
            text += "Здесь даже паук не селится — пусто."
        else:
            folders = [i for i in items if i['mimeType'] == 'application/vnd.google-apps.folder']
            files = [i for i in items if i['mimeType'] != 'application/vnd.google-apps.folder']
            if folders:
                text += "<b>Подпапки:</b>"
                for f in sorted(folders, key=lambda x: x['name'].lower()):
                    text += f"📁 <code>{f['name']}/</code>"
                text += ""
            if files:
                text += "<b>Файлы:</b>"
                for f in sorted(files, key=lambda x: x['name'].lower()):
                    size = f" ({f['size']} байт)" if f.get('size') else ""
                    text += f"📄 <code>{f['name']}</code>{size}"
        await update.message.reply_text(text, parse_mode='HTML')
    except Exception as e:
        logger.error(f"❌ Ошибка /path: {e}")
        await update.message.reply_text(
            get_message('search_error')
        )

async def reload_lists(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Перезагрузка чёрного и белого списков.
    Args:
        update (Update): Объект обновления от Telegram
        context (ContextTypes.DEFAULT_TYPE): Контекст обработчика
    """
    if not update.message or not update.effective_user:
        return
    user = update.effective_user
    # Проверяем, является ли пользователь администратором
    if not user.username or user.username.lower() not in {u.lower() for u in ALLOWED_USERS}:
        await update.message.reply_text("❌ Доступ запрещён.")
        return
    if not access_manager:
        await update.message.reply_text("❌ Система доступа не инициализирована.")
        return
    # Обновляем списки
    access_manager.update_lists()
    await update.message.reply_text(
        f"✅ Списки успешно перезагружены.\n"
        f"Белый список: {len(access_manager.whitelist)} пользователей\n"
        f"Чёрный список: {len(access_manager.blacklist)} пользователей\n"
    )
    logger.info(f"🔄 Администратор {user.username} перезагрузил списки доступа.")

# --- Команда /reset_bans ---
async def reset_bans(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Сброс лимитов для пользователя или всех пользователей (только для админов).
    Args:
        update (Update): Объект обновления от Telegram
        context (ContextTypes.DEFAULT_TYPE): Контекст обработчика
    """
    if not update.message or not update.effective_user:
        return
    user = update.effective_user
    # Проверяем, является ли пользователь администратором
    if not user.username or user.username.lower() not in {u.lower() for u in ALLOWED_USERS}:
        await update.message.reply_text(get_message('admin_only'))
        return
    # Получаем параметры команды
    args = context.args
    if not args:
        await update.message.reply_text(
            "Использование: /reset_bans <имя_пользователя или 'all'>"
        )
        return
    target = args[0].lower()
    if target == 'all':
        # Сбросить все лимиты
        user_activity.clear()
        banned_users.clear()
        user_ban_start_times.clear()
        user_ban_times.clear()
        await update.message.reply_text(get_message('reset_all_success'))
        logger.info(f"🔄 Администратор {user.username} сбросил все лимиты")
    else:
        # Сбросить лимиты для конкретного пользователя
        username = target.lstrip('@')  # Убираем @ если есть
        reset_user_limits(username)
        await update.message.reply_text(
            get_message('reset_success', username=username)
        )
        logger.info(f"🔄 Администратор {user.username} сбросил лимиты для пользователя {username}")

# --- Класс для работы с Google Drive файлами ---
class FileManager:
    """
    Класс для работы с файлами в Google Drive.
    Предоставляет методы для поиска, скачивания и получения информации о файлах.
    """
    def __init__(self, drive):
        """
        Инициализация менеджера файлов.
        Args:
            drive: Сервис Google Drive
        """
        self.drive = drive

    def find_folder(self, parent_id: str, name: str) -> Optional[str]:
        """
        Ищет папку по имени в заданной родительской папке.
        Args:
            parent_id (str): ID родительской папки
            name (str): Имя папки для поиска
        Returns:
            Optional[str]: ID найденной папки или None
        """
        # Формируем запрос к API Google Drive
        query = f"mimeType='application/vnd.google-apps.folder' and name='{name}' and '{parent_id}' in parents and trashed=false"
        try:
            res = self.drive.files().list(q=query, fields="files(id)").execute()
            folder_id = res['files'][0]['id'] if res['files'] else None
            if folder_id:
                logger.info(f"🔍 Найдена папка: '{name}' (ID: {folder_id})")
            else:
                logger.debug(f"📁 Папка не найдена: '{name}' в родителе {parent_id}")
            return folder_id
        except Exception as e:
            logger.error(f"❌ Ошибка поиска папки '{name}': {e}")
            return None

    def find_file(self, folder_id: str, filename: str) -> Optional[str]:
        """
        Ищет файл по имени в заданной папке.
        Args:
            folder_id (str): ID папки
            filename (str): Имя файла для поиска
        Returns:
            Optional[str]: ID найденного файла или None
        """
        # Формируем запрос к API Google Drive
        query = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
        try:
            res = self.drive.files().list(q=query, fields="files(id)").execute()
            file_id = res['files'][0]['id'] if res['files'] else None
            if file_id:
                logger.info(f"📎 Найден файл: '{filename}' (ID: {file_id})")
            else:
                logger.debug(f"📄 Файл не найден: '{filename}' в папке {folder_id}")
            return file_id
        except Exception as e:
            logger.error(f"❌ Ошибка поиска файла '{filename}': {e}")
            return None

    def get_file_modified_time(self, file_id: str) -> Optional[datetime]:
        """
        Получает время модификации файла.
        Args:
            file_id (str): ID файла
        Returns:
            Optional[datetime]: Время модификации файла или None
        """
        try:
            # Получаем информацию о файле
            info = self.drive.files().get(fileId=file_id, fields="modifiedTime").execute()
            t = info['modifiedTime']
            # Преобразуем строку времени в объект datetime
            dt = datetime.strptime(t, "%Y-%m-%dT%H:%M:%S.%fZ")
            # Применяем смещение часового пояса
            dt_with_tz = dt.replace(tzinfo=timezone.utc) + timedelta(hours=TIMEZONE_OFFSET)
            return dt_with_tz
        except Exception as e:
            logger.error(f"❌ Ошибка получения времени файла {file_id}: {e}")
            return None

    def download_file(self, file_id: str, local_path: str) -> bool:
        """
        Скачивает файл из Google Drive в локальную директорию.
        Args:
            file_id (str): ID файла в Google Drive
            local_path (str): Локальный путь для сохранения файла
        Returns:
            bool: True, если файл успешно скачан, False в противном случае
        """
        try:
            # Получаем медиа-поток файла
            request = self.drive.files().get_media(fileId=file_id)
            with open(local_path, 'wb') as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                # Скачиваем файл по частям
                while not done:
                    status, done = downloader.next_chunk()
            logger.info(f"✅ Файл успешно скачан: ID={file_id}, путь={local_path}")
            return True
        except Exception as e:
            logger.error(f"❌ Ошибка при скачивании файла ID={file_id} в {local_path}: {e}")
            return False

    def list_files_in_folder(self, folder_id: str, max_results: int = 100) -> List[Dict]:
        """
        Получает список файлов и папок в заданной папке.
        Args:
            folder_id (str): ID папки
            max_results (int): Максимальное количество результатов
        Returns:
            List[Dict]: Список файлов и папок
        """
        try:
            # Формируем запрос к API
            query = f"'{folder_id}' in parents and trashed=false"
            res = self.drive.files().list(q=query, pageSize=max_results, fields="files(id, name, mimeType, size)").execute()
            return res.get('files', [])
        except Exception as e:
            logger.error(f"❌ Ошибка списка файлов в папке {folder_id}: {e}")
            return []

    def check_write_permission(self, file_id: str) -> bool:
        """
        Проверяет, есть ли у учетных данных бота права на редактирование файла.
        Args:
            file_id (str): ID файла в Google Drive
        Returns:
            bool: True, если есть права на запись, False в противном случае
        """
        try:
            # Получаем информацию о файле, включая разрешения
            info = self.drive.files().get(fileId=file_id, fields="capabilities/canEdit, permissions").execute()
            can_edit = info.get('capabilities', {}).get('canEdit', False)
            logger.debug(f"Проверка прав на запись для файла {file_id}: canEdit={can_edit}")
            return can_edit
        except Exception as e:
            logger.error(f"❌ Ошибка проверки прав на запись для файла {file_id}: {e}")
            return False

    def update_list_file(self, file_id: str, usernames: List[str]) -> bool:
        """
        Обновляет содержимое текстового файла в Google Drive.
        Args:
            file_id (str): ID файла в Google Drive
            usernames (List[str]): Список username для записи
        Returns:
            bool: True, если успешно, False в противном случае
        """
        try:
            # 1. Сначала получаем метаданные файла, чтобы узнать его MIME-тип
            file_metadata = self.drive.files().get(fileId=file_id, fields="mimeType, name").execute()
            mime_type = file_metadata.get('mimeType', 'text/plain')
            filename = file_metadata.get('name', 'list.txt')

            # 2. Создаем новый контент
            content = "\n".join([f"@{u}" for u in usernames]) + "\n" # Каждый юзер с новой строки, с @
            media_body = MediaIoBaseUpload(io.BytesIO(content.encode('utf-8')), mimetype=mime_type, resumable=True)

            # 3. Обновляем файл
            updated_file = self.drive.files().update(
                fileId=file_id,
                media_body=media_body
            ).execute()

            logger.info(f"✅ Файл списка {filename} (ID: {file_id}) успешно обновлён. Новое содержимое: {usernames}")
            return True
        except Exception as e:
            logger.error(f"❌ Ошибка обновления файла списка {file_id}: {e}")
            return False

# --- Класс для поиска данных в Excel ---
class LocalDataSearcher:
    """
    Класс для поиска данных в локальных Excel файлах.
    Предоставляет методы для асинхронного поиска по серийным номерам.
    """
    @staticmethod
    async def search_by_number_async(filepath: str, number: str) -> List[str]:
        """
        Асинхронный поиск терминала по серийному номеру в Excel файле.
        Args:
            filepath (str): Путь к Excel файлу
            number (str): Серийный номер для поиска
        Returns:
            List[str]: Список результатов поиска
        """
        loop = asyncio.get_event_loop()
        # Выполняем синхронную операцию в пуле потоков
        return await loop.run_in_executor(executor, LocalDataSearcher._search_by_number_sync, filepath, number)
    @staticmethod

    def _search_by_number_sync(filepath: str, number: str) -> List[str]:
        """
        Синхронная реализация поиска терминала по серийному номеру.
        Args:
            filepath (str): Путь к Excel файлу
            number (str): Серийный номер для поиска
        Returns:
            List[str]: Список результатов поиска
        """
        number_upper = number.strip().upper()
        results = []
        try:
            # Логирование запроса
            logger.info(f"🔍 Поиск терминала по СН: {number_upper}")
            # Проверка существования файла
            if not os.path.exists(filepath):
                logger.error(f"❌ Файл не существует: {filepath}")
                return results
            # Открываем Excel файл
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet = wb["Терминалы"] if "Терминалы" in wb.sheetnames else None
            if not sheet:
                logger.warning(f"⚠️ Лист 'Терминалы' не найден в {filepath}")
                wb.close()
                return results
            # Проверка наличия данных в файле
            if sheet.max_row < 2:
                logger.warning(f"⚠️ Файл {filepath} пуст или не содержит данных")
                wb.close()
                return results
            found = False
            # Проходим по строкам таблицы
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 17 or not row[5]:  # СН в столбце F (индекс 5)
                    continue
                # Извлечение данных
                sn = str(row[5]).strip().upper()
                if sn != number_upper:
                    continue
                found = True
                equipment_type = str(row[4]).strip() if row[4] else "Не указано"
                model = str(row[6]).strip() if row[6] else "Не указано"
                request_num = str(row[7]).strip() if row[7] else "Не указано"
                status = str(row[8]).strip() if row[8] else "Не указано"
                storage = str(row[13]).strip() if row[13] else "Не указано"
                issue_status = str(row[14]).strip() if row[14] else ""
                engineer = str(row[15]).strip() if row[15] else "Не указано"
                issue_date = str(row[16]).strip() if row[16] else "Не указано"
                # Регистронезависимые проверки
                status_lower = status.lower()
                issue_status_lower = issue_status.lower()
                # Формируем базовые поля
                response_parts = [
                    f"<b>СН:</b> <code>{sn}</code>/n",
                    f"<b>Тип оборудования:</b> <code>{equipment_type}</code>/n",
                    f"<b>Модель терминала:</b> <code>{model}</code>/n",
                ]
                # --- Логика по статусу ---
                if status_lower == "на складе":
                    response_parts.append(f"<b>Статус оборудования:</b> <code>{status}</code>/n")
                    response_parts.append(f"<b>Место на складе:</b> <code>{storage}</code>/n")
                elif status_lower in ["не работоспособно", "выведено из эксплуатации"]:
                    response_parts.append(f"<b>Статус оборудования:</b> <code>{status}</code> — как труп в багажнике/n")
                    response_parts.append(f"<b>Место на складе:</b> <code>{storage}</code> — можно разобрать на запчасти/n")
                elif status_lower == "зарезервировано":
                    response_parts.append(f"<b>Статус оборудования:</b> <code>{status}</code>/n")
                    response_parts.append(f"<b>Место на складе:</b> <code>{storage}</code>/n")
                    if issue_status_lower == "выдан":
                        # Показываем всё: место, инженера, дату
                        response_parts.append(f"<b>Заявка:</b> <code>{request_num}</code>/n")
                        response_parts.append(f"<b>Выдан инженеру:</b> <code>{engineer}</code>/n")
                        response_parts.append(f"<b>Дата выдачи:</b> <code>{issue_date}</code>/n")
                    # Если не выдан — ничего больше не добавляем
                else:
                    # Все остальные статусы: просто показываем статус
                    response_parts.append(f"<b>Статус оборудования:</b> <code>{status}</code>/n")
                    # Можно добавить место, если нужно, но по ТЗ — не требуется
                # Формируем итоговый текст
                header = "ℹ️ <b>Информация о терминале</b>/n"
                result_text = header + "" + "".join(response_parts)
                results.append(result_text)
            wb.close()
            # Логирование результата поиска
            if found:
                logger.info(f"✅ Найден терминал по СН: {number_upper}")
            else:
                logger.info(f"❌ Терминал не найден по СН: {number_upper}")
        except openpyxl.utils.exceptions.InvalidFileException as e:
            logger.error(f"❌ Ошибка чтения Excel (поврежденный файл): {filepath} - {e}")
        except openpyxl.utils.exceptions.IllegalCharacterError as e:
            logger.error(f"❌ Ошибка чтения Excel (недопустимые символы): {filepath} - {e}")
        except Exception as e:
            logger.error(f"❌ Неожиданная ошибка при чтении Excel {filepath}: {e}", exc_info=True)
        return results

async def handle_search(update: Update, query: str, user=None, username=None):
    """
    Общая логика поиска терминала по серийному номеру.
    Args:
        update (Update): Объект обновления от Telegram
        query (str): Запрос пользователя
        user (User, optional): Объект пользователя (если известен)
        username (str, optional): Username пользователя (если известен)
    """
    # Определяем пользователя, если не передан
    if user is None:
        user = update.effective_user
    if username is None:
        username = user.username if user.username else str(user.id)

    # Проверяем доступ в приватном чате
    if update.message.chat.type == 'private':
        if not user.username or not access_manager.is_allowed(user.username.lower()):
            await update.message.reply_text(
                get_message('access_denied')
            )
            return

    # Проверяем лимиты DDoS
    if not check_user_limit(username):
        # Получаем время до разблокировки
        ban_start = user_ban_start_times.get(username)
        ban_time = user_ban_times.get(username, 10)
        if ban_start:
            remaining_time = ban_start + timedelta(minutes=ban_time) - (datetime.now(timezone.utc) + timedelta(hours=TIMEZONE_OFFSET))
            minutes_left = int(remaining_time.total_seconds() // 60)
            await update.message.reply_text(
                f"Стопэ! Ты слишком быстро пишешь! Тебе нужно немного передышки.\n"
                f"Абажди {minutes_left} минут и попробуй снова.",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                get_message('ddos_blocked'),
                parse_mode='HTML'
            )
        return

    # Извлекаем серийные номера (разделенные запятой)
    numbers = [extract_number(num_str) for num_str in query.split(',')]
    numbers = [num for num in numbers if num]  # Убираем пустые значения
    if not numbers:
        await update.message.reply_text(
            get_message('invalid_number'),
            parse_mode='HTML'
        )
        return

    # Отправляем промежуточное сообщение только один раз
    try:
        if len(numbers) == 1:
            await update.message.reply_text(
                get_message('search_start', number=numbers[0]),
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                f"🔍 Копаю в архивах... Где-то были эти СН: {', '.join(numbers)}...",
                parse_mode='HTML'
            )
    except Exception as e:
        logger.error(f"❌ Не удалось отправить статус-сообщение: {e}")
        return

    global LAST_FILE_ID, LAST_FILE_DATE, LAST_FILE_DRIVE_TIME, LAST_FILE_LOCAL_PATH
    # Проверка: есть ли загруженный файл
    if not LAST_FILE_ID or not LAST_FILE_LOCAL_PATH:
        logger.warning("❌ Нет данных: файл не был предзагружен при старте.")
        try:
            await update.message.reply_text(
                get_message('no_file')
            )
        except Exception as e:
            logger.error(f"❌ Не удалось отправить ответ об отсутствии файла: {e}")
        return
    if not os.path.exists(LAST_FILE_LOCAL_PATH):
        logger.warning(f"❌ Локальный файл не найден: {LAST_FILE_LOCAL_PATH}")
        try:
            await update.message.reply_text(
                get_message('file_not_found_local')
            )
        except Exception as e:
            logger.error(f"❌ Ошибка отправки сообщения: {e}")
        return

    # Получаем актуальное время файла в Google Drive
    try:
        gs = GoogleServices()
        fm = FileManager(gs.drive)
        current_drive_time = fm.get_file_modified_time(LAST_FILE_ID)
        if not current_drive_time:
            logger.warning(f"⚠️ Не удалось получить время изменения файла: {LAST_FILE_ID}")
            # Продолжаем с кэшированным временем
        else:
            # Проверяем, нужно ли обновить
            local_time = datetime.fromtimestamp(os.path.getmtime(LAST_FILE_LOCAL_PATH), tz=timezone.utc)
            if LAST_FILE_DRIVE_TIME is None or current_drive_time > LAST_FILE_DRIVE_TIME:
                logger.info(f"🔄 Файл в облаке новее ({current_drive_time.isoformat()} > {LAST_FILE_DRIVE_TIME}). Скачивание...")
                try:
                    if fm.download_file(LAST_FILE_ID, LAST_FILE_LOCAL_PATH):
                        LAST_FILE_DRIVE_TIME = current_drive_time
                        logger.info(f"✅ Файл обновлён: {LAST_FILE_LOCAL_PATH}")
                    else:
                        logger.error("❌ Не удалось скачать обновлённый файл. Используем старую версию.")
                        try:
                            await update.message.reply_text(
                                get_message('file_update_error')
                            )
                        except Exception as e:
                            logger.error(f"❌ Ошибка отправки предупреждения: {e}")
                except Exception as e:
                    logger.error(f"❌ Ошибка при скачивании файла: {e}", exc_info=True)
                    try:
                        await update.message.reply_text(
                            get_message('file_update_success')
                        )
                    except Exception as e_inner:
                        logger.error(f"❌ Ошибка отправки уведомления: {e_inner}")
    except Exception as e:
        logger.error(f"❌ Критическая ошибка при проверке обновления файла: {e}", exc_info=True)
        try:
            await update.message.reply_text(
                get_message('search_error')
            )
        except Exception as e_inner:
            logger.error(f"❌ Ошибка отправки сообщения: {e_inner}")

    # Поиск по локальному файлу
    try:
        # Используем асинхронный поиск для каждого номера
        lds = LocalDataSearcher()
        all_results = []
        for number in numbers:
            results = await lds.search_by_number_async(LAST_FILE_LOCAL_PATH, number)
            all_results.extend(results)
        if not all_results:
            if len(numbers) == 1:
                await update.message.reply_text(
                    get_message('no_terminal', number=numbers[0]),
                    parse_mode='HTML'
                )
            else:
                await update.message.reply_text(
                    f"Терминалы с СН {', '.join(numbers)} не найдены.",
                    parse_mode='HTML'
                )
            return

        # Отправляем результаты по одному
        for result in all_results:
            try:
                if len(result) > 4096:
                    truncated = result[:4050] + "<i>... (обрезано)</i>"
                    await update.message.reply_text(truncated, parse_mode='HTML')
                else:
                    await update.message.reply_text(result, parse_mode='HTML')
            except Exception as e:
                logger.error(f"❌ Ошибка отправки результата: {e}")
                try:
                    await update.message.reply_text(
                        "Нашёл терминал, но не могу показать — что-то сломалось./n"
                        "Попробуй позже или скажи админу."
                    )
                except Exception as e_inner:
                    logger.error(f"❌ Ошибка отправки fallback-сообщения: {e_inner}")
    except Exception as e:
        logger.error(f"❌ Ошибка при поиске в Excel: {e}", exc_info=True)
        try:
            await update.message.reply_text(
                get_message('search_error')
            )
        except Exception as e_inner:
            logger.error(f"❌ Ошибка отправки сообщения об ошибке чтения: {e_inner}")

# Обработчик команды /refresh ---
async def refresh_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Принудительное обновление файла с Google Drive (только для админов).
    Args:
        update (Update): Объект обновления от Telegram
        context (ContextTypes.DEFAULT_TYPE): Контекст обработчика
    """
    if not update.message or not update.effective_user:
        return
    user = update.effective_user
    # Проверяем, является ли пользователь администратором
    if not user.username or user.username.lower() not in {u.lower() for u in ALLOWED_USERS}:
        await update.message.reply_text("❌ Доступ запрещён.")
        return
    global LAST_FILE_ID, LAST_FILE_DATE, LAST_FILE_DRIVE_TIME, LAST_FILE_LOCAL_PATH
    # Проверяем наличие данных о файле
    if not LAST_FILE_ID or not LAST_FILE_LOCAL_PATH:
        await update.message.reply_text("❌ Нет данных о файле для обновления.")
        return
    try:
        await update.message.reply_text("🔄 Обновление файла с Google Drive...")
        gs = GoogleServices()
        fm = FileManager(gs.drive)
        # Получаем текущее время файла в Google Drive
        current_drive_time = fm.get_file_modified_time(LAST_FILE_ID)
        if not current_drive_time:
            await update.message.reply_text("❌ Не удалось получить время изменения файла.")
            return
        # Скачиваем файл
        if fm.download_file(LAST_FILE_ID, LAST_FILE_LOCAL_PATH):
            LAST_FILE_DRIVE_TIME = current_drive_time
            await update.message.reply_text(
                f"✅ Файл успешно обновлён!\n"
                f"Дата изменения: {current_drive_time.strftime('%d.%m.%Y %H:%M:%S')}"
            )
            logger.info(f"🔄 Файл обновлён администратором {user.username}")
        else:
            await update.message.reply_text("❌ Не удалось обновить файл.")
    except Exception as e:
        logger.error(f"❌ Ошибка при обновлении файла: {e}")
        await update.message.reply_text("❌ Произошла ошибка при обновлении файла.")

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обработка сообщений: только команды и упоминания в чатах.
    Args:
        update (Update): Объект обновления от Telegram
        context (ContextTypes.DEFAULT_TYPE): Контекст обработчика
    """
    if not update.message or not update.message.text:
        return

    text = update.message.text.strip()
    bot_username = context.bot.username.lower() if context.bot.username else ""
    chat_type = update.message.chat.type
    user = update.effective_user

    #logger.info(f"DEBUG: bot_username = '{bot_username}'")
    #logger.info(f"DEBUG: text = '{text}'")
    #logger.info(f"DEBUG: chat_type = '{chat_type}'")

    # Проверяем доступ в приватном чате
    if chat_type == 'private':
        if not user.username or not access_manager.is_allowed(user.username.lower()):
            await update.message.reply_text(get_message('access_denied'))
            return

        # Проверяем лимиты DDoS
        username = user.username if user.username else str(user.id)
        if not check_user_limit(username):
            # Получаем время до разблокировки
            ban_start = user_ban_start_times.get(username)
            ban_time = user_ban_times.get(username, 10)
            if ban_start:
                remaining_time = ban_start + timedelta(minutes=ban_time) - (datetime.now(timezone.utc) + timedelta(hours=TIMEZONE_OFFSET))
                minutes_left = int(remaining_time.total_seconds() // 60)
                await update.message.reply_text(
                    f"Ты слишком быстро пишешь! Тебе нужно немного передышки.\n"
                    f"Пожалуйста, подожди {minutes_left} минут и попробуй снова.",
                    parse_mode='HTML'
                )
            else:
                await update.message.reply_text(get_message('ddos_blocked'), parse_mode='HTML')
            return

        # Обработка команды /s
        if text.startswith("/s"):
            query = text[2:].strip()
            if not query:
                await update.message.reply_text(get_message('missing_number'), parse_mode='HTML')
                return
            await handle_search(update, query, user, username)
            return

        # Обработка других команд
        elif text.startswith('/'):
            await update.message.reply_text(get_message('unknown_command'), parse_mode='HTML')
        else:
            # Отправляем помощь для обычных сообщений
            await update.message.reply_text(get_message('help'), parse_mode='HTML')
        return

    # В групповых чатах (group/supergroup) — только команды и упоминания
    if chat_type in ['group', 'supergroup']:
        # Проверяем лимиты DDoS
        username = user.username if user.username else str(user.id)
        if not check_user_limit(username):
            # Получаем время до разблокировки
            ban_start = user_ban_start_times.get(username)
            ban_time = user_ban_times.get(username, 10)
            if ban_start:
                remaining_time = ban_start + timedelta(minutes=ban_time) - (datetime.now(timezone.utc) + timedelta(hours=TIMEZONE_OFFSET))
                minutes_left = int(remaining_time.total_seconds() // 60)
                await update.message.reply_text(
                    f"Ты слишком быстро пишешь! Тебе нужно немного передышки.\n"
                    f"Пожалуйста, подожди {minutes_left} минут и попробуй снова.",
                    parse_mode='HTML'
                )
            else:
                await update.message.reply_text(get_message('ddos_blocked'), parse_mode='HTML')
            return

        # --- ИСПРАВЛЕННАЯ ЛОГИКА ---
        # 1. Обработка команды /s в группе (например, /s 123456)
        if text.startswith("/s"):
            query = text[2:].strip()
            if not query:
                await update.message.reply_text(get_message('missing_number'), parse_mode='HTML')
                return
            await handle_search(update, query, user, username)
            return # Завершаем обработку после команды

        # 2. Обработка упоминания бота в группе (например, @Sklad_bot 123456)
        #    Это должно быть вне условия text.startswith("/s")
        mention_match = re.search(rf'@{re.escape(bot_username)}\s+(.+)', text, re.IGNORECASE)
        if mention_match:
            query = mention_match.group(1).strip()
            if not query:
                await update.message.reply_text(
                    "Укажи серийный номер после упоминания бота.\n"
                    f"Пример: @{context.bot.username} AB123456",
                    parse_mode='HTML'
                )
                return
            # --- ИСПРАВЛЕНИЕ ---
            # Извлекаем username из user объекта (уже есть выше)
            # username = user.username if user.username else str(user.id) # Уже определено
            await handle_search(update, query, user, username) # передаем username
            return # Завершаем обработку после упоминания

        # Все остальные сообщения в группе — игнорируем
        return

    # Для каналов (channel) — только упоминания (если бот добавлен как админ)
    if chat_type == 'channel':
        # Проверяем упоминание: @Sklad_bot ...
        username = user.username if user.username else str(user.id)
        mention_match = re.search(rf'@{re.escape(bot_username)}\s+(.+)', text, re.IGNORECASE)
        if mention_match:
            query = mention_match.group(1).strip()
            if not query:
                # Отправка сообщений в каналы может быть ограничена
                logger.warning("Попытка ответить в канале на пустой запрос. Это может не сработать.")
                # Можно попробовать reply_text, но часто не работает. Лучше логировать.
                # await update.message.reply_text(...) # Можем не иметь права отвечать
                logger.info("Получено упоминание в канале с пустым запросом.")
                return
            await handle_search(update, query, user, username) # передаем username
            return
        # Все остальные сообщения в канале — игнорируем
        return

def main():
    """
    Основная функция запуска бота.
    Инициализирует конфигурацию, создает обработчики и запускает бота.
    """
    try:
        init_config()
    except Exception as e:
        logger.critical(f"❌ Критическая ошибка: {e}")
        return

    # Создаем приложение Telegram бота
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    # Инициализация AccessManager
    global access_manager
    gs = GoogleServices()
    access_manager = AccessManager(gs.drive)
    access_manager.update_lists()

    # Предзагружаем последний файл
    preload_latest_file()

    # Регистрируем функцию для удаления временного файла при выходе
    atexit.register(lambda: os.remove("temp_google_creds.json") if os.path.exists("temp_google_creds.json") else None)

    # Добавляем обработчики команд
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("path", show_path))
    app.add_handler(CommandHandler("reload_lists", reload_lists))
    app.add_handler(CommandHandler("restart", restart_bot))
    app.add_handler(CommandHandler("refresh", refresh_file))
    app.add_handler(CommandHandler("reset_bans", reset_bans))
    app.add_handler(CommandHandler("whitelist", manage_whitelist))
    app.add_handler(CommandHandler("blacklist", manage_blacklist))

    # Добавляем обработчик сообщений
    app.add_handler(MessageHandler(filters.TEXT, handle_message))

    logger.info("🚀 Бот запущен. Готов к работе.")
    app.run_polling()

if __name__ == '__main__':
    main()